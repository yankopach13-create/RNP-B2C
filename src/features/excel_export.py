"""Сборка Excel-отчёта РНП B2C (лист на каждый блок)."""

from __future__ import annotations

from dataclasses import dataclass, field
from io import BytesIO
import re

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config.constants import EXCLUDED_RNP_GROUPS
from data.loaders import AppData
from features.client_segments import compute_segment_client_metrics
from features.clients import (
    COL_CUMULATIVE,
    COL_METRIC,
    COL_WEEK,
    _build_metric_rows,
    _clients_bk_week_count,
    _compute_client_metrics,
    _prepare_checks_clients,
)
from features.data_prep import filter_sales_by_report_week
from features.excise_liquid import WeekCalculationConfig
from features.focus import build_focus_display_df
from features.fill_free_products import build_fill_free_table
from features.hookah_products import build_hookah_products_table
from features.lfl import build_lfl_factor_table
from features.metrics import (
    _build_category_sales_general_rows,
    _build_category_sales_group_rows,
    _build_financial_b2c_rows,
    _build_financial_group_rows,
    _build_shop_economy_table,
    _build_turnover_summary,
    _can_build_category_sales,
    _can_build_financial_metrics,
    _financial_row_style_kind,
    _financial_rows_to_dataframe,
)

_EXCLUDED_GROUPS_NORM = frozenset(g.casefold() for g in EXCLUDED_RNP_GROUPS)


def _drop_excluded_groups(df: pd.DataFrame | None) -> pd.DataFrame | None:
    """Убирает строки исключённых групп (например «ИКС») из выгрузки Excel."""
    if df is None or df.empty or "Группа" not in df.columns:
        return df
    norm = df["Группа"].astype(str).str.strip().str.casefold()
    return df.loc[~norm.isin(_EXCLUDED_GROUPS_NORM)].copy()


def _filter_groups_order_list(groups_order: list[str] | None) -> list[str] | None:
    if not groups_order:
        return groups_order
    return [
        g
        for g in groups_order
        if str(g).strip().casefold() not in _EXCLUDED_GROUPS_NORM
    ]


HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
GROUP_FONT = Font(bold=True, size=11)
BOLD_FONT = Font(bold=True, size=11)
NORMAL_FONT = Font(size=11)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")


@dataclass
class ExcelSheetSpec:
    name: str
    table: pd.DataFrame | CategoriesSheetParts
    row_kinds: list[str] = field(default_factory=list)
    layout: str = "table"  # table | categories


@dataclass
class CategoriesSheetParts:
    general: pd.DataFrame | None = None
    turnover: pd.DataFrame | None = None
    groups: pd.DataFrame | None = None
    general_row_kinds: list[str] = field(default_factory=list)
    groups_row_kinds: list[str] = field(default_factory=list)


def rnp_b2c_excel_filename(report_week: int | None) -> str:
    if report_week is not None:
        return f"РНП B2C {report_week}.xlsx"
    return "РНП B2C.xlsx"


def build_rnp_b2c_excel_bytes(
    data: AppData,
    prepared,
    week_config: WeekCalculationConfig | None,
) -> bytes | None:
    """Возвращает содержимое xlsx или None, если нет листов для экспорта."""
    sheets = collect_rnp_b2c_sheets(data, prepared, week_config)
    if not sheets:
        return None

    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for spec in sheets:
            sheet_name = _sanitize_sheet_name(spec.name)
            if spec.layout == "categories" and isinstance(spec.table, CategoriesSheetParts):
                _write_categories_sheet(writer, sheet_name, spec.table)
            else:
                export_df = _prepare_table_for_excel(spec.table)
                export_df.to_excel(
                    writer,
                    sheet_name=sheet_name,
                    index=False,
                )
                ws = writer.sheets[sheet_name]
                _style_worksheet(
                    ws,
                    export_df,
                    row_kinds=spec.row_kinds,
                    sheet_name=spec.name,
                )

    return buffer.getvalue()


def collect_rnp_b2c_sheets(
    data: AppData,
    prepared,
    week_config: WeekCalculationConfig | None,
) -> list[ExcelSheetSpec]:
    sheets: list[ExcelSheetSpec] = []
    report_week = week_config.report_week if week_config else None
    lfl_week = week_config.lfl_week if week_config else None
    excise_report = week_config.excise_liquid_report if week_config else 0.0
    excise_lfl = week_config.excise_liquid_lfl if week_config else 0.0

    df = prepared.df if prepared is not None else None
    df_report = _drop_excluded_groups(_report_sales_df(df, data.sales, week_config))
    groups_order = _filter_groups_order_list(data.groups_order_rnp)

    if df_report is not None and _can_build_financial_metrics(df_report):
        finance_parts: list[tuple[pd.DataFrame, list[str]]] = []
        b2c_rows = _build_financial_b2c_rows(
            df_report,
            data.client_segments,
            report_week,
            excise_liquid_report_qty=excise_report,
        )
        finance_parts.append(_financial_rows_to_dataframe(b2c_rows))

        group_rows = _build_financial_group_rows(df_report, groups_order)
        if group_rows:
            finance_parts.append(_financial_rows_to_dataframe(group_rows))

        finance_table, finance_kinds = _stack_tables_with_styles(finance_parts)
        if finance_table is not None:
            sheets.append(
                ExcelSheetSpec(
                    name="Финансы",
                    table=finance_table,
                    row_kinds=finance_kinds,
                )
            )

    if df_report is not None and _can_build_category_sales(df_report):
        cat_general = None
        general_rows = _build_category_sales_general_rows(
            df_report, data.category_order_rnp
        )
        general_kinds: list[str] = []
        if general_rows:
            cat_general, general_kinds = _financial_rows_to_dataframe(general_rows)
            cat_general = _prepare_table_for_excel(cat_general)

        turnover_table = _build_turnover_summary(
            data.turnover_90,
            data.turnover_week,
            data.categories,
            data.category_order_rnp,
        )
        turnover_prepared = _prepare_turnover_for_excel(turnover_table)

        cat_groups = None
        group_kinds: list[str] = []
        cat_group_rows = _build_category_sales_group_rows(
            df_report,
            data.category_order_rnp,
            groups_order,
        )
        if cat_group_rows:
            cat_groups, group_styles = _financial_rows_to_dataframe(cat_group_rows)
            group_kinds = group_styles
            cat_groups = _prepare_table_for_excel(cat_groups)

        parts = CategoriesSheetParts(
            general=cat_general,
            turnover=turnover_prepared,
            groups=cat_groups,
            general_row_kinds=general_kinds,
            groups_row_kinds=group_kinds,
        )
        if any((parts.general is not None, parts.turnover is not None, parts.groups is not None)):
            sheets.append(
                ExcelSheetSpec(
                    name="Продажи категорий",
                    table=parts,
                    layout="categories",
                )
            )

    if df_report is not None:
        focus_table = build_focus_display_df(df_report, data.focus)
        if focus_table is not None and not focus_table.empty:
            sheets.append(
                ExcelSheetSpec(
                    name="Фокус",
                    table=_prepare_table_for_excel(focus_table),
                )
            )

    client_table = _build_client_block_export(
        data.checks_clients,
        data.client_segments,
        report_week,
    )
    if client_table is not None:
        sheets.append(
            ExcelSheetSpec(
                name="Клиентский блок",
                table=_prepare_table_for_excel(client_table),
            )
        )

    if df_report is not None:
        shop_table = _build_shop_economy_table(df_report, data.shops_order)
        if shop_table is not None and not shop_table.empty:
            sheets.append(
                ExcelSheetSpec(
                    name="Экономика магазинов",
                    table=_prepare_table_for_excel(shop_table),
                )
            )

    lfl_table = build_lfl_factor_table(
        data.lfl,
        data.categories,
        lfl_week,
        report_week,
        data.category_order_rnp,
        excise_liquid_lfl_qty=excise_lfl,
        excise_liquid_report_qty=excise_report,
    )
    if lfl_table is not None and not lfl_table.empty:
        sheets.append(
            ExcelSheetSpec(
                name="Факторный анализ",
                table=_prepare_table_for_excel(lfl_table),
            )
        )

    if df_report is not None or data.focus_hookah is not None:
        hookah_table = build_hookah_products_table(
            df_report,
            data.focus_hookah,
            data.groups,
            report_week,
        )
        if hookah_table is not None and not hookah_table.empty:
            sheets.append(
                ExcelSheetSpec(
                    name="Кальянная продукция",
                    table=_prepare_table_for_excel(hookah_table),
                )
            )

    if data.focus_fill_free is not None:
        fill_free_table, _ = build_fill_free_table(
            data.focus_fill_free,
            data.groups,
            report_week,
        )
        if fill_free_table is not None and not fill_free_table.empty:
            sheets.append(
                ExcelSheetSpec(
                    name="Fill free",
                    table=_prepare_table_for_excel(fill_free_table),
                )
            )

    return sheets


def _write_categories_sheet(
    writer: pd.ExcelWriter,
    sheet_name: str,
    parts: CategoriesSheetParts,
) -> None:
    """Общие слева, оборачиваемость справа; подразделения ниже — без служебных колонок."""
    written = False
    top_rows = 0

    if parts.general is not None and not parts.general.empty:
        parts.general.to_excel(
            writer,
            sheet_name=sheet_name,
            index=False,
            startrow=0,
            startcol=0,
        )
        top_rows = len(parts.general)
        left_cols = len(parts.general.columns)
        written = True
    else:
        left_cols = 0

    if parts.turnover is not None and not parts.turnover.empty:
        start_col = left_cols + 1 if left_cols else 0
        parts.turnover.to_excel(
            writer,
            sheet_name=sheet_name,
            index=False,
            startrow=0,
            startcol=start_col,
        )
        top_rows = max(top_rows, len(parts.turnover))
        written = True

    bottom_start = top_rows + 2 if top_rows else 0
    if parts.groups is not None and not parts.groups.empty:
        parts.groups.to_excel(
            writer,
            sheet_name=sheet_name,
            index=False,
            startrow=bottom_start,
            startcol=0,
        )
        written = True

    if not written:
        return

    ws = writer.sheets[sheet_name]
    if parts.general is not None and not parts.general.empty:
        _style_financial_block(
            ws,
            parts.general,
            start_row=1,
            start_col=1,
            row_kinds=parts.general_row_kinds or ["normal"] * len(parts.general),
        )
    if parts.turnover is not None and not parts.turnover.empty:
        turnover_col = (left_cols + 2) if left_cols else 1
        _style_turnover_block(ws, parts.turnover, start_row=1, start_col=turnover_col)
    if parts.groups is not None and not parts.groups.empty:
        _style_financial_block(
            ws,
            parts.groups,
            start_row=bottom_start + 1,
            start_col=1,
            row_kinds=parts.groups_row_kinds or ["normal"] * len(parts.groups),
        )

    _autofit_worksheet(ws)


def _style_worksheet(
    ws,
    df: pd.DataFrame,
    *,
    row_kinds: list[str] | None = None,
    sheet_name: str,
) -> None:
    if df.empty:
        return

    ncols = len(df.columns)
    nrows = len(df)

    for col_idx in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    kinds = row_kinds or ["normal"] * nrows
    for row_idx in range(nrows):
        kind = kinds[row_idx] if row_idx < len(kinds) else "normal"
        excel_row = row_idx + 2
        if kind == "spacer":
            continue

        for col_idx in range(1, ncols + 1):
            cell = ws.cell(row=excel_row, column=col_idx)
            value = df.iloc[row_idx, col_idx - 1]
            cell.value = value
            cell.border = THIN_BORDER

            header = str(df.columns[col_idx - 1])
            if col_idx == ncols or header in (
                "Значение",
                "Продажи, шт.",
                "Продажи с НДС",
                "Накопительно",
            ) or header.startswith("Неделя "):
                cell.alignment = RIGHT
            else:
                cell.alignment = LEFT

            if sheet_name == "Фокус" and col_idx == 1 and str(value).strip():
                cell.font = GROUP_FONT
            elif sheet_name in ("Финансы", "Продажи категорий") and col_idx == 1:
                group_val = str(df.iloc[row_idx, 0]).strip()
                metric_val = str(df.iloc[row_idx, 1]).strip() if ncols > 1 else ""
                if group_val and metric_val:
                    if col_idx == 1:
                        cell.font = GROUP_FONT
                elif kind == "weight" and col_idx == 2:
                    cell.font = BOLD_FONT
                elif kind == "weight" and col_idx == 3:
                    cell.font = BOLD_FONT
                else:
                    cell.font = NORMAL_FONT
            elif sheet_name in (
                "Клиентский блок",
                "Кальянная продукция",
                "Fill free",
            ) and col_idx == 1 and str(value).strip():
                cell.font = BOLD_FONT
            else:
                cell.font = NORMAL_FONT

    _autofit_worksheet(ws)


def _style_financial_block(
    ws,
    df: pd.DataFrame,
    *,
    start_row: int,
    start_col: int,
    row_kinds: list[str],
) -> None:
    ncols = len(df.columns)
    for col_offset in range(ncols):
        cell = ws.cell(row=start_row, column=start_col + col_offset)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        cell.value = df.columns[col_offset]

    for row_idx in range(len(df)):
        kind = row_kinds[row_idx] if row_idx < len(row_kinds) else "normal"
        excel_row = start_row + 1 + row_idx
        if kind == "spacer":
            continue

        for col_offset in range(ncols):
            cell = ws.cell(row=excel_row, column=start_col + col_offset)
            value = df.iloc[row_idx, col_offset]
            cell.value = value
            cell.border = THIN_BORDER
            col_name = df.columns[col_offset]

            if col_offset == ncols - 1 or col_name == "Значение":
                cell.alignment = RIGHT
            else:
                cell.alignment = LEFT

            group_val = str(df.iloc[row_idx, 0]).strip()
            metric_val = str(df.iloc[row_idx, 1]).strip() if ncols > 1 else ""
            if col_offset == 0 and group_val and metric_val:
                cell.font = GROUP_FONT
            elif kind == "weight" and col_offset == 1:
                cell.font = BOLD_FONT
            elif kind == "weight" and col_offset == 2:
                cell.font = BOLD_FONT
            else:
                cell.font = NORMAL_FONT


def _style_turnover_block(
    ws,
    df: pd.DataFrame,
    *,
    start_row: int,
    start_col: int,
) -> None:
    for col_offset, col_name in enumerate(df.columns):
        cell = ws.cell(row=start_row, column=start_col + col_offset)
        cell.value = col_name
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    for row_idx in range(len(df)):
        excel_row = start_row + 1 + row_idx
        for col_offset in range(len(df.columns)):
            cell = ws.cell(row=excel_row, column=start_col + col_offset)
            cell.value = df.iloc[row_idx, col_offset]
            cell.border = THIN_BORDER
            if col_offset == 0:
                cell.alignment = LEFT
                cell.font = NORMAL_FONT
            else:
                cell.alignment = RIGHT
                cell.font = NORMAL_FONT


def _autofit_worksheet(ws) -> None:
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 42)


def _stack_tables_with_styles(
    parts: list[tuple[pd.DataFrame, list[str]]],
    spacer_rows: int = 2,
) -> tuple[pd.DataFrame | None, list[str]]:
    prepared: list[tuple[pd.DataFrame, list[str]]] = []
    for table, table_kinds in parts:
        if table is None or table.empty:
            continue
        prepared.append((_prepare_table_for_excel(table), table_kinds))

    if not prepared:
        return None, []

    all_columns = list(
        dict.fromkeys(col for table, _ in prepared for col in table.columns)
    )
    frames: list[pd.DataFrame] = []
    kinds: list[str] = []
    for idx, (table, table_kinds) in enumerate(prepared):
        if idx > 0 and spacer_rows > 0:
            spacer = pd.DataFrame(
                [[""] * len(all_columns) for _ in range(spacer_rows)],
                columns=all_columns,
            )
            frames.append(spacer)
            kinds.extend(["spacer"] * spacer_rows)
        frames.append(table.reindex(columns=all_columns))
        kinds.extend(table_kinds)

    return pd.concat(frames, ignore_index=True), kinds


def _prepare_turnover_for_excel(
    turnover_table: pd.DataFrame | None,
) -> pd.DataFrame | None:
    if turnover_table is None or turnover_table.empty:
        return None
    out = turnover_table.copy()
    out.index.name = "Категория"
    return _prepare_table_for_excel(out.reset_index())


def _prepare_table_for_excel(df: pd.DataFrame | None) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    out = df.copy()
    if not isinstance(out.index, pd.RangeIndex):
        out = out.reset_index()
        first_col = out.columns[0]
        if first_col in ("index", "level_0"):
            index_name = df.index.name or "Категория"
            out = out.rename(columns={first_col: index_name})
    out = out.loc[:, [c for c in out.columns if not str(c).startswith("_gap")]]
    return out


def _report_sales_df(
    prepared_df: pd.DataFrame | None,
    sales_raw: pd.DataFrame | None,
    week_config: WeekCalculationConfig | None,
) -> pd.DataFrame | None:
    if prepared_df is not None:
        if week_config is not None:
            return filter_sales_by_report_week(prepared_df, week_config.report_week)
        return prepared_df
    if sales_raw is not None:
        if week_config is not None:
            return filter_sales_by_report_week(sales_raw, week_config.report_week)
        return sales_raw
    return None


def _build_client_block_export(
    checks_clients: pd.DataFrame | None,
    client_segments: pd.DataFrame | None,
    report_week: int | None,
) -> pd.DataFrame | None:
    if checks_clients is None or checks_clients.empty:
        return None
    try:
        df = _prepare_checks_clients(checks_clients)
    except ValueError:
        return None

    if report_week is None:
        report_week = int(df[COL_WEEK].max())

    week_label = f"Неделя {report_week}"
    week_df = df[df[COL_WEEK] == report_week]
    metrics = _compute_client_metrics(df, week_df)
    metrics.update(
        compute_segment_client_metrics(
            client_segments,
            report_week,
            _clients_bk_week_count(week_df),
        )
    )
    rows = _build_metric_rows(metrics)
    return pd.DataFrame(rows, columns=[COL_METRIC, COL_CUMULATIVE, week_label])


def _sanitize_sheet_name(name: str) -> str:
    cleaned = re.sub(r"[\[\]\:\*\?\/\\]", " ", str(name)).strip()
    return (cleaned or "Лист")[:31]
